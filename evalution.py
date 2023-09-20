import openpyxl
#uploading responses file
df1=openpyxl.load_workbook("studentrespones.xlsx")
sheet = df1['Form Responses 1']
df2=openpyxl.load_workbook("key.xlsx")
sheet2 = df2['key']
count = 0
k=0
#instalizing empty list
data=[]
data1=[]
data2=[]
data3=[]
data4=[]
data5=[]
data6=[]
data7=[]
data8=[]
data9=[]
data10=[]
data11=[]
#MATHS
#maths set A

for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='A':
        for j in range(8,38):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data.insert(i,count)
        count = 0
#print(data)
        
import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 98

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='A':
    #print(data[u])
    sheet1.cell(i,column_index).value = data[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data)

#maths setB
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='B':
        for j in range(8,38):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data1.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 98

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='B':
    #print(data[u])
    sheet1.cell(i,column_index).value = data1[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data1)
#maths setC
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='C':
        for j in range(8,38):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data2.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 98

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='C':
    #print(data[u])
    sheet1.cell(i, column_index).value = data2[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data2)
#PHY
#phy set A
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='A':
        for j in range(38,68):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data3.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 99

  #col= (row - 1) % 1+1
  
u=0
#PHY 
#phy set A 
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='A':
    #print(data[u])
    sheet1.cell(i, column_index).value = data3[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data3)
#phy set B
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='B':
        for j in range(38,68):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data4.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 99

  #col= (row - 1) % 1+1
  
u=0
#phy set B  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='B':
    #print(data[u])
    sheet1.cell(i, column_index).value = data4[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data4)
#phy set C
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='C':
        for j in range(38,68):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data5.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 99

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='C':
    #print(data[u])
    sheet1.cell(i, column_index).value = data5[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data5)
#CHE
#che set A
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='A':
        for j in range(68,98):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data6.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 100

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='A':
    #print(data[u])
    sheet1.cell(i, column_index).value = data6[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data6)
# che set B
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='B':
        for j in range(68,98):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data7.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 100

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='B':
    #print(data[u])
    sheet1.cell(i, column_index).value = data7[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data7)
#CHE set C
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='C':
        for j in range(68,98):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data8.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 100

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='C':
    #print(data[u])
    sheet1.cell(i, column_index).value = data8[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data8)
#TOTAL
# set A
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='A':
        for j in range(8,98):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data9.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 101

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='A':
    #print(data[u])
    sheet1.cell(i, column_index).value = data9[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data9)
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='B':
        for j in range(8,98):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data10.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 101

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='B':
    #print(data[u])
    sheet1.cell(i, column_index).value = data10[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data10)


#tatal set C
for i in range(2,sheet.max_row+1):
    if sheet.cell(i,7).value=='C':
        for j in range(8,98):
            k=j-7
            if sheet.cell(i,j).value=="None":
                pass
                #print("sheet: " ,i ,j," -" , sheet2.cell(i,j).value)
            elif sheet.cell(i,j).value== sheet2.cell(5,k).value:
                count += 4
            else:
                count -=1
        data11.insert(i,count)
        count = 0

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Select the sheet you want to insert data into
# Define the list of data
#data = []
sheet1 = workbook['Form Responses 1']

# Insert the data into the sheet
column_index = 101

  #col= (row - 1) % 1+1
  
u=0
  
for i in range(2,sheet.max_row+1):  
  if sheet.cell(i,7).value=='C':
    #print(data[u])
    sheet1.cell(i, column_index).value = data11[u]
    u=u+1
    
workbook.save('studentrespones.xlsx')
print(data11)
# ranking process

import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('studentrespones.xlsx')

# Get the sheet you want to fetch data from
sheet = workbook['Form Responses 1']

# Create an empty list to store the column data
column_data = []

# Get the maximum number of rows in the sheet
max_row = sheet.max_row

# Iterate through each row of the sheet and append the value of the desired column to the list
for row in range(2, max_row+1):
    cell_value = sheet.cell(row=row, column=101).value # Column 8 is column "H" in Excel
    column_data.append(cell_value)

# Print the list
print(column_data)
#count=len(column_data)
#print(count)

student_marks=[]
student_marks = column_data
sorted_marks = sorted(student_marks, reverse=True)

ranked_list = [0] * len(student_marks)  # create an empty list of the same length as student_marks

for i, marks in enumerate(sorted_marks):
    rank = i + 1
    for j, student in enumerate(student_marks):
        if marks == student and ranked_list[j] == 0:  # check if the rank for the student has already been assigned
            ranked_list[j] = rank

print(ranked_list)

import openpyxl
workbook = openpyxl.load_workbook('studentrespones.xlsx')
sheet = workbook['Form Responses 1']
data = ranked_list
column = 'CX'
row_num = 2
for item in data:    
    cell = f'{column}{row_num}'      
    sheet[cell] = item   
    row_num += 1
workbook.save('studentrespones.xlsx')
